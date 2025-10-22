from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from APP.DTO.solicitacao_carga_dto import User
from APP.Config.ihs_config import _ensure_driver
from APP.Core.solicitacao_carga_core import Path
from openpyxl import load_workbook, Workbook
from tkinter import messagebox
from typing import Dict, List
from datetime import datetime
from time import sleep, time
import tkinter as tk
import pandas as pd
import threading
import logging
import random
import os

_driver = None
_lock = threading.Lock()


ARQ = r'\\172.17.67.14\findev$\Automação - CNH\SENHAS IHS.xlsx'

def preencher_campo(driver, value, texto):
    """
    Preenche um campo de texto em um formulário web.

    Parâmetros:
        driver (webdriver): Instância do Selenium WebDriver.
        value (tuple): Localizador do elemento (By, valor).
        texto (str): Texto a ser inserido no campo.

    Retorno:
        None
    """
    campo = driver.find_element(By.CSS_SELECTOR, value)
    campo.click()
    espera_personalizada(
        lambda: campo.send_keys(texto),
        inicio_random=2,
        fim_random=4
    )


def clicar_em_um_botao(driver, value):
    """
    Clica em um botão simulando a tecla ENTER.

    Parâmetros:
        driver (webdriver): Instância do Selenium WebDriver.
        value (tuple): Localizador do botão (By, valor).

    Retorno:
        None
    """
    espera_personalizada(
        lambda: driver.find_element(By.CSS_SELECTOR, value).send_keys(Keys.ENTER),
        inicio_random=2,
        fim_random=4
    )


def clica_na_aba(wdw, css_selector, texto_ancora):
    """
    Clica em uma aba específica com base no texto exibido.

    Parâmetros:
        driver (webdriver): Instância do Selenium WebDriver.
        css_selector (str): Seletor CSS para localizar as abas.
        texto_ancora (str): Texto da aba desejada (case-insensitive).

    Retorno:
        None
    """
    aba = wdw.until(
        lambda d: next(
            (
                el for el in d.find_elements(By.CSS_SELECTOR, css_selector)
                if texto_ancora in el.text.strip().lower()
            ),
            None
        )
    )

    if aba:
        aba.click()
    else:
        print("❌ Aba com texto não encontrada.")

def clica_em_um_botao_pelo_nome(driver, value, texto_ancora):
    """
    Clica em uma aba específica com base no texto exibido.

    Parâmetros:
        driver (webdriver): Instância do Selenium WebDriver.
        css_selector (str): Seletor CSS para localizar as abas.
        texto_ancora (str): Texto da aba desejada (case-insensitive).

    Retorno:
        None
    """
    elementos = driver.find_elements(By.CSS_SELECTOR, value)
    for a in elementos:
        if a.text.lower() == texto_ancora:
            espera_personalizada(
                lambda: a.click()
            )
            break


def clicar_pelo_atributo(driver, atributo, texto_comparar, value):
    """
    Clica em um elemento com base no valor de um atributo.

    Parâmetros:
        driver (webdriver): Instância do Selenium WebDriver.
        atributo (str): Nome do atributo HTML a ser comparado.
        texto_comparar (str): Valor esperado do atributo.
        value (tuple): Localizador do grupo de elementos (By, valor).

    Retorno:
        None
    """
    entradas = driver.find_elements(By.CSS_SELECTOR, value)
    for i in entradas:
        if i.get_attribute(atributo).strip().lower() == texto_comparar:
            espera_personalizada(
                lambda: i.click()
            )
            break



def cria_indicador_de_tempo_execucao(
    start_time: float,
    end_time: float,
    nome_robo: str,
    arquivo: str = r"\\172.17.67.14\findev$\indicadores.xlsx",
    aba: str = "Execuções",
):
    """
    Cria ou atualiza um arquivo Excel com indicadores de tempo de execução.

    Parâmetros:
        start_time (float): Timestamp do início.
        end_time (float): Timestamp do fim.
        nome_robo (str): Nome do robô que executou.
        arquivo (str): Caminho completo do arquivo Excel.
        aba (str): Nome da aba onde salvar os dados.

    Retorno:
        str: Mensagem de sucesso.
        Exception: PermissionError caso o arquivo esteja aberto.
    """
    # garante que a pasta existe
    os.makedirs(os.path.dirname(arquivo), exist_ok=True)

    # abre ou cria o arquivo Excel
    if os.path.exists(arquivo):
        wb = load_workbook(arquivo)
        if aba in wb.sheetnames:
            ws = wb[aba]
        else:
            ws = wb.create_sheet(aba)
            ws.append(["Hora Inicial", "Hora Final", "Total Execução (min)", "Robô"])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = aba
        ws.append(["Hora Inicial", "Hora Final", "Total Execução (min)", "Robô"])

    # converte timestamps em horas legíveis
    hora_inicial = datetime.fromtimestamp(start_time).strftime("%Y-%m-%d %H:%M:%S")
    hora_final = datetime.fromtimestamp(end_time).strftime("%Y-%m-%d %H:%M:%S")
    tempo_total_min = round((end_time - start_time) / 60, 2)
    tempo_total_min = f'00:{tempo_total_min}'.replace('.', ':')

    # grava uma linha
    ws.append([hora_inicial, hora_final, tempo_total_min, nome_robo])

    try:
        wb.save(arquivo)
        return "Indicador salvo com sucesso!"
    except PermissionError:
        raise PermissionError(f"Feche o arquivo '{arquivo}' no Excel e tente novamente.")


def espera_personalizada(*acoes, retorno=False, inicio_random=3, fim_random=8):
    """
    Executa uma ou mais ações com tentativas repetidas até que tenham sucesso. Caso tenha retorno, adicione a função que deve receber retorno como a última.

    Parâmetros:
        *acoes (callable): Funções ou lambdas que executam ações.
        **retorno (bool): Variável boolena para verificar se é necessário retornar algo.

    Retorno:
        None
    """
    contador = 0
    while True:
        try:
            # espera aleatória para simular comportamento humano
            sleep(random.randint(inicio_random, fim_random))

            # executa todas menos a última
            for acao in acoes[:-1] if retorno else acoes:
                acao()

            # se tiver retorno, executa a última e devolve
            if retorno:
                return acoes[-1]()
            
            break
        except Exception:
            if contador == 5:
                break
            contador += 1
            continue


def sair_ihs(driver, value):
    driver.switch_to.default_content()
    driver.find_element(By.CSS_SELECTOR, value).click()

def clica_em_prosseguir(driver, value):
    driver.find_element(By.CSS_SELECTOR, value).click()

def get_all_users(lojas):
    """
    Carrega todos os usuários definidos no arquivo `.env` e cria objetos `User`.

    Parâmetros:
        None

    Retorno:
        list[User]: Lista de instâncias de usuários com os atributos
        `codigo`, `usuario`, `senha` e `nome_loja`.
    """
    usuarios = []
    prefixos = tela(lojas)

    # Validação simples
    tamanhos = {k: len(v) for k, v in prefixos.items()}
    if len(set(tamanhos.values())) != 1:
        raise ValueError(f"Listas com tamanhos diferentes: {tamanhos}")

    for l, c, u, s in zip(prefixos['LOJAS'], prefixos['CODIGOS'], prefixos['USUARIOS'], prefixos['SENHAS']):
        codigo = c
        usuario = u
        senha = s
        nome_loja = l

        if codigo and usuario and senha:
            usuarios.append(User(codigo=codigo, usuario=usuario, senha=senha, nome_loja=nome_loja))
    return usuarios

def _norm_loja(nome: str) -> str:
    return str(nome).strip().upper().replace(" ", "-")

def _normaliza_lista_lojas_param(param: str) -> List[str]:
    """
    Converte o parâmetro 'lojas' em uma lista de keys normalizadas.
    Ex.: 'Ares Motos, Cajazeiras' -> ['ares-motos', 'cajazeiras'] (exemplo)
    """
    partes = [p.strip() for p in param.split(",") if p.strip()]
    return [_norm_loja(p) for p in partes]

def tela(lojas: str) -> Dict[str, List[str]]:
    """
    Lê o Excel e retorna o dicionário de listas filtrado:
    {'LOJAS': [...], 'CODIGOS': [...], 'USUARIOS': [...], 'SENHAS': [...]}

    Parâmetros:
        lojas (str): 'all' ou lista separada por vírgula (ordem preservada).
                     Aceita tanto nome exibido quanto key normalizada.

    Retorno:
        dict[str, list[str]]
    """
    # 1) Lê o Excel
    df = pd.read_excel(ARQ, dtype=str)  # dtype=str evita NaN -> float
    esperadas = {"LOJAS", "CODIGOS", "USUARIOS", "SENHAS"}
    if not esperadas.issubset(set(df.columns)):
        raise ValueError(f"Colunas esperadas: {esperadas} | Encontradas: {set(df.columns)}")

    # 2) Cria a coluna de chave normalizada
    df["_LOJA_KEY"] = df["LOJAS"].map(_norm_loja)

    # 3) Determina quais keys usar e a ordem
    if lojas.lower() == "all":
        # usa TODAS as lojas na ordem do arquivo (removendo duplicatas por primeira ocorrência)
        # se houver duplicatas, consideramos a primeira ocorrência
        ordem_keys = []
        vistos = set()
        for k in df["_LOJA_KEY"]:
            if k not in vistos:
                ordem_keys.append(k)
                vistos.add(k)
    else:
        # usa SOMENTE as lojas passadas na URL, preservando a ordem fornecida
        ordem_keys = _normaliza_lista_lojas_param(lojas)

    if not ordem_keys:
        raise ValueError("Nenhuma loja informada/derivada do parâmetro.")

    # 4) Monta listas NA ORDEM selecionada, pegando a primeira linha quando houver duplicidade
    df_idx = df.set_index("_LOJA_KEY")
    lojas_out, codigos, usuarios, senhas = [], [], [], []

    lojas_inexistentes = []
    for key in ordem_keys:
        if key not in df_idx.index:
            lojas_inexistentes.append(key)
            continue

        row = df_idx.loc[key]
        if isinstance(row, pd.DataFrame):
            row = row.iloc[0]  # pega a primeira ocorrência

        lojas_out.append(_norm_loja(row["LOJAS"]))        # mantém padronização usada no resto do fluxo
        codigos.append(str(row["CODIGOS"]).strip())
        usuarios.append(str(row["USUARIOS"]).strip())
        senhas.append(str(row["SENHAS"]).strip())

    if lojas_inexistentes:
        # falha explícita para evitar executar sem todas as lojas pedidas
        raise ValueError(f"Lojas não encontradas no Excel (use nomes válidos ou equivalentes normalizados): {lojas_inexistentes}")

    return {"LOJAS": lojas_out, "CODIGOS": codigos, "USUARIOS": usuarios, "SENHAS": senhas}

def solicitacao_carga_main(lojas):
        # Data atual formatada
    hoje = datetime.today().strftime("%Y-%m-%d")

    # Caminho base
    path_file_log = r'\\172.17.67.14\findev$\Automação - CNH\Solicitação de carga\Logs\\'
    arquivo_log = f'logs_solicitacao_carga-{hoje}.log'.replace('/', '-').replace(':', '_')
    caminho = os.path.join(path_file_log, arquivo_log)

    # Configuração do logging
    logging.basicConfig(
        filename=caminho,
        filemode='a',
        level=logging.ERROR,
        format="%(asctime)s - %(message)s",
        datefmt="%d/%m/%Y %H:%M:%S"
    )

    # Marca o início da execução (para métricas)
    start_time = time()

    path = Path()

    driver, wdw, PASTA_DOWNLOADS = _ensure_driver(_driver, _lock)

    driver.get(path.Url.url)

    # Busca todos os usuários cadastrados
    try:
        usuarios = get_all_users(lojas)
    except Exception as e:
        return False, f'Erro ao buscar os usuários.\nDescrição: {str(e)}'

    for user in usuarios:


        try:
            # ===============================
            #             LOGIN
            # ===============================
            espera_personalizada(inicio_random=1,fim_random=3)
            driver.find_element(By.CSS_SELECTOR, path.Login.campo_code).send_keys(user.codigo)
            espera_personalizada(inicio_random=1,fim_random=3)
            driver.find_element(By.CSS_SELECTOR, path.Login.campo_user).send_keys(user.usuario)
            espera_personalizada(inicio_random=1,fim_random=3)
            driver.find_element(By.CSS_SELECTOR, path.Login.campo_password).send_keys(user.senha)
            espera_personalizada(inicio_random=1,fim_random=3)

            submit = wdw.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, path.Login.btn_entrar))
            )
            submit.click()
            
            prosseguir = wdw.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, path.Login.btn_prosseguir))
            )

            prosseguir.click()

            # ===============================
            #       NAVEGAÇÃO NO MENU
            # ===============================
            espera_personalizada()

            try:
                btn_mensagem = WebDriverWait(driver, timeout=15).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, path.MenuPrincipal.btn_mensagem))
                )

                btn_mensagem.click()
            except:
                pass
            
            clica_na_aba(wdw, path.MenuPrincipal.aba_consorcio, 'consórcio')  # Consórcio
            espera_personalizada(inicio_random=1, fim_random=3)
            clica_na_aba(wdw, path.MenuPrincipal.aba_formularios_download, 'formulários e download')

            espera_personalizada(inicio_random=1, fim_random=3)
            clica_na_aba(wdw, path.MenuPrincipal.aba_consorcio_side_menu, 'consórcio - formulários/download')

            espera_personalizada(inicio_random=1, fim_random=3)
            clica_na_aba(wdw, path.MenuPrincipal.aba_solicitacao_carga, 'solicitação carga de arquivos')

            # ===============================
            #        ACESSO AO IFRAME
            # ===============================
            espera_personalizada(
                lambda: driver.switch_to.default_content(),
                lambda: driver.switch_to.frame(0),
                inicio_random=2,
                fim_random=4
            )

            check_box = wdw.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, path.Frame.checkbox_ccc))
            )

            check_box.click()

            clicar_pelo_atributo(driver, 'value', 'solicitar carga', path.Frame.btn_solicitar_carga)

            # ===============================
            #        ACESSO AO ALERT
            # ===============================
            alert = espera_personalizada(
                lambda: driver.switch_to.alert,
                retorno=True
            )
            alert.accept()

            # ===============================
            #      VOLTA PARA O IFRAME
            # ===============================
            espera_personalizada(
                lambda: driver.switch_to.default_content(),
                lambda: driver.switch_to.frame(0),
                inicio_random=2,
                fim_random=4
            )

            # ===============================
            #   ADICIONANDO A DATA DE HOJE
            # ===============================
            hoje = datetime.now().strftime("%d/%m/%y")

            sleep(random.randint(2, 4))

            preencher_campo(driver, path.Frame.entry_a, hoje)
            preencher_campo(driver, path.Frame.entry_de, hoje)

            sleep(random.randint(2, 4))
            clicar_pelo_atributo(driver, 'value', 'confirmar', path.Frame.btn_confirmar)

            # ===============================
            #        ACESSO AO ALERT
            # ===============================
            alert = espera_personalizada(
                lambda: driver.switch_to.alert,
                retorno=True
            )
            alert.accept()

            sleep(random.randint(2, 4))
            # ===============================
            #          SAIR DA LOJA
            # ===============================
            sair_ihs(driver, path.Logout.btn_sair)

            logging.error(msg=f'✅ Solicitação efetuada na loja: {user.nome_loja}.\n{'-'*60}\n')

        except Exception as e:
            logging.error(msg=f'❌ Erro ao solicitar a carga da loja {user.nome_loja}.\nDescrição: {str(e)}.\n{'-'*60}\n')
            driver.get(path.Url.url)
            continue
    
    return True, 'Solicitação de Carga realizada, confira os arquivos, para saber se ele criou tudo corretamente.'
