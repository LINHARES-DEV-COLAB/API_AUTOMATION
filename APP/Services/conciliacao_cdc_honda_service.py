from selenium.webdriver.support import expected_conditions as EC
from APP.Config.ihs_config import PASTA_DOWNLOADS, _ensure_driver, busca_dados_db
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from typing import Dict, List, Tuple, Optional
from openpyxl import load_workbook, Workbook
from selenium.webdriver.common.by import By
from datetime import datetime, timedelta
from APP.DTO.ihs_dto import User
from collections import Counter
from unidecode import unidecode
from tkinter import messagebox
from dotenv import load_dotenv
from pathlib import Path as P
from time import sleep, time
from fuzzywuzzy import fuzz
import unicodedata
import logging
import random
import pandas
import os
import re

def verificar_similaridade_fuzzy(nome1, nome2, limite=90):
    """
    Usa fuzzy matching para verificar similaridade
    """
    similaridade = fuzz.ratio(nome1.lower(), nome2.lower())
    return similaridade >= limite, similaridade


def preencher_campo(driver, locator, texto):
    """
    Preenche um campo de texto em um formulário web.

    Parâmetros:
        driver (webdriver): Instância do Selenium WebDriver.
        locator (tuple): Localizador do elemento (By, valor).
        texto (str): Texto a ser inserido no campo.

    Retorno:
        None
    """
    campo = driver.find_element(*locator)
    campo.click()
    espera_personalizada(
        lambda: campo.send_keys(texto),
        inicio_random=2,
        fim_random=4
    )


def clicar_em_um_botao(driver, locator):
    """
    Clica em um botão simulando a tecla ENTER.

    Parâmetros:
        driver (webdriver): Instância do Selenium WebDriver.
        locator (tuple): Localizador do botão (By, valor).

    Retorno:
        None
    """
    espera_personalizada(
        lambda: driver.find_element(*locator).send_keys(Keys.ENTER),
        inicio_random=2,
        fim_random=4
    )


def clica_na_aba(wdw, css_selector, texto_ancora):
    """
    Clica em uma aba específica com base no texto exibido.

    Parâmetros:
        driver (webdriver): Instância do Selenium WebDriver.
        css_selector (str): Seletor CSS para localizar as abas.
        texto_ancora (str): Texto da aba desejada (case-insensitive).

    Retorno:
        None
    """
    aba = wdw.until(
        lambda d: next(
            (
                el for el in d.find_elements(By.CSS_SELECTOR, css_selector)
                if texto_ancora in el.text.strip().lower()
            ),
            None
        )
    )

    if aba:
        aba.click()
    else:
        print("❌ Aba com texto não encontrada.")


def clicar_pelo_atributo(driver, atributo, texto_comparar, locator):
    """
    Clica em um elemento com base no valor de um atributo.

    Parâmetros:
        driver (webdriver): Instância do Selenium WebDriver.
        atributo (str): Nome do atributo HTML a ser comparado.
        texto_comparar (str): Valor esperado do atributo.
        locator (tuple): Localizador do grupo de elementos (By, valor).

    Retorno:
        None
    """
    entradas = driver.find_elements(*locator)
    for i in entradas:
        if i.get_attribute(atributo).lower() == texto_comparar:
            espera_personalizada(
                lambda: i.click()
            )
            break

def tenta_passar_pagina_atras(driver):
    try:
        espera_personalizada(
            lambda: driver.find_element(By.ID, 'IMAGE2').click(),
            inicio_random=6,
            fim_random=10
        )
        return True
    except:
        return False

def tenta_passar_pagina(driver):
    try:
        espera_personalizada(
            lambda: driver.find_element(By.ID, 'IMAGE1').click(),
            inicio_random=6,
            fim_random=10
        )
        return True
    except:
        return False

def pega_valores_liquidos(driver, wdw, EC):
    try:
        abas = driver.window_handles
        driver.switch_to.window(abas[-1])
        driver.set_window_position(-3566,-3566)
        driver.set_window_size(1366, 768)
        tabela = wdw.until(EC.presence_of_element_located((By.ID, 'Grid1ContainerTbl')))

        lista_nomes = []
        lista_valor = []

        contador = 1
        while contador != 0:
            try:
                id_linha_ajustado = ajusta_id('#Grid1ContainerRow', contador)
                linha = tabela.find_element(By.CSS_SELECTOR, id_linha_ajustado)

                # pega o nome e o valor em colunas específicas
                lista_nomes.append(unidecode(linha.find_elements(By.CSS_SELECTOR, '.gx-attribute')[2].text.strip().upper()))
                lista_valor.append(linha.find_elements(By.CSS_SELECTOR, '.gx-attribute')[-1].text.strip())
                contador += 1
            except:
                contador = 0

        driver.close()
        abas = driver.window_handles
        driver.switch_to.window(abas[0])
        driver.switch_to.frame(0)

        return (lista_nomes, lista_valor)
    except Exception as e:
        return (False, False)

def ajusta_id(primeira_parte, numero_lote):
    """
    Ajusta o ID de um elemento conforme o número do lote.

    Parâmetros:
        primeira_parte (str): Prefixo do ID.
        numero_lote (int): Número do lote.

    Retorno:
        str: ID ajustado conforme o tamanho do número.
    """
    if numero_lote < 10:
        return f'{primeira_parte}_000{numero_lote}'
    elif numero_lote > 9 and numero_lote < 100:
        return f'{primeira_parte}_00{numero_lote}'
    elif numero_lote > 99 and numero_lote < 1000:
        return f'{primeira_parte}_0{numero_lote}'
    elif numero_lote > 999 and numero_lote < 10000:
        return f'{primeira_parte}_{numero_lote}'


def parse_valor(valor_str: str) -> float:
    """
    Converte um número no formato brasileiro (1.234,56) para float.

    Parâmetros:
        valor_str (str): String com o valor monetário.

    Retorno:
        float: Valor convertido.
    """
    valor_str = valor_str.strip()
    if not valor_str:
        return 0.0
    return float(valor_str.replace('.', '').replace(',', '.'))

def _valor_aproximado(a, b, tol_abs: float = 50.0, tol_pct: Optional[float] = None) -> Tuple[bool, float]:
    """
    Retorna (ok, diff).
    ok=True se |a-b| <= tol_abs OU (se tol_pct for definido) |a-b| <= base*tol_pct.
    """
    a = float(a)
    b = float(b)
    diff = abs(a - b)
    ok = diff <= tol_abs
    if tol_pct is not None:
        base = max(abs(a), abs(b), 1e-9)
        ok = ok or (diff <= base * tol_pct)
    return ok, diff

def compara_valores(
    nome_db: str,
    valor_db,
    dados_honda: Dict[str, float],
    empresa: str,
    titulo,
    *,
    tol_abs: float = 50.0,
    tol_pct: Optional[float] = None,
    thr_exato: int = 92,
    thr_parcial: int = 80
) -> Tuple[bool, str, float]:
    """
    - Usa fuzzy matching para achar o melhor nome em dados_honda.
    - Compara o valor com tolerância (abs e/ou percentual).
    - Evita KeyError, pois usa a chave realmente parecida (melhor_nome).
    """
    alvo = nome_db.strip().upper()
    melhor_nome = None
    melhor_score = -1

    # Loop para achar o nome mais parecido
    for nome in dados_honda:
        eh_similar, percentual = verificar_similaridade_fuzzy(alvo, nome.strip().upper(), limite=thr_exato)
        # Guarda sempre o melhor score, mesmo que não bata o limite
        if percentual > melhor_score:
            melhor_score = percentual
            melhor_nome = nome

    if melhor_nome is None:
        return False, f"❌ Nenhum cliente correspondente encontrado. Empresa: {empresa}", valor_db

    try:
        valor_honda = float(dados_honda[melhor_nome])
        valor_db = float(valor_db)
    except KeyError:
        return False, f'❌ O Cliente {melhor_nome} não é encontrado nos dados da Honda.', valor_db

    # Lógica de decisão
    if melhor_score >= thr_exato:
        # Match forte — compara valores com tolerância
        ok, diff = _valor_aproximado(valor_db, valor_honda, tol_abs, tol_pct)
        if ok:
            if valor_db == valor_honda:
                # valores idênticos
                return True, (
                    f"✅ Cliente '{melhor_nome}' (match com {nome_db} de {melhor_score}%) "
                    f"com título {titulo}: valor ok ({valor_db}). Empresa: {empresa}"
                ), valor_db
            else:
                # dentro da tolerância → normaliza
                tipo_ajuste = "acréscimo" if valor_honda > valor_db else "decréscimo"
                diferenca = abs(valor_honda - valor_db)
                valor_normalizado = valor_db
                return True, (
                    f"⚠️ Cliente '{melhor_nome}' (match com {nome_db} de {melhor_score}%) com título {titulo}: "
                    f"valor precisa ser ajustado por {tipo_ajuste} de R$ {diferenca:.2f}. "
                    f"Empresa: {empresa}"
                ), valor_normalizado
        else:
            return True, (
                f"❌ Cliente '{melhor_nome}' (match com {nome_db} de {melhor_score}%) "
                f"com título {titulo}: valores incompatíveis — Linx: {valor_db} | Honda: {valor_honda} "
                f"(diferença {diff:.2f}). Empresa: {empresa}"
            ), valor_db

    elif thr_parcial <= melhor_score < thr_exato:
        return False, (
            f"❌ Nome incompatível: Linx ('{nome_db}') x Honda ('{melhor_nome}') "
            f"com título {titulo}: valores incompatíveis — Linx: {valor_db} | Honda: {valor_honda} "
            f"(match {melhor_score}%). Empresa: {empresa}"
        ), valor_db

    else:
        return False, False, valor_db


def cria_arquivo_log(loja: str, logs: list,
                     pasta: str = r'\\172.17.67.14\findev$\Automação - CDC\Documentos_Logs'):
    """
    Cria um arquivo de log para uma loja específica.

    Parâmetros:
        loja (str): Nome ou identificador da loja.
        logs (list): Lista de mensagens de log.
        pasta (str): Caminho da pasta onde salvar o arquivo (default configurado).

    Retorno:
        None
    """
    hoje = datetime.now().strftime("%d-%m-%Y")
    os.makedirs(pasta, exist_ok=True)

    header = '-' * 40
    linhas = [str(l) for l in logs if l]  # evita None e garante string
    texto = f'{header}\nLoja: {loja}\n' + '\n'.join(linhas) + f'\n{header}\n\n'

    path_file = os.path.join(pasta, f'log_loja_{loja}_{hoje}.txt')
    # utf-8-sig garante visualização correta até no Bloco de Notas
    with open(path_file, 'w', encoding='utf-8-sig', newline='\n') as f:
        f.write(texto)
    
    path_file = os.path.join(r'\\fileserver\tic\Processos\Pauline - Processos\Honda - CDC\Logs', f'log_loja_{loja}_{hoje}.txt')
    with open(path_file, 'w', encoding='utf-8-sig', newline='\n') as f:
        f.write(texto)


def cria_indicador_de_tempo_execucao(
    start_time: float,
    end_time: float,
    nome_robo: str,
    arquivo: str = r"\\172.17.67.14\findev$\indicadores.xlsx",
    aba: str = "Execuções",
):
    """
    Cria ou atualiza um arquivo Excel com indicadores de tempo de execução.

    Parâmetros:
        start_time (float): Timestamp do início.
        end_time (float): Timestamp do fim.
        nome_robo (str): Nome do robô que executou.
        arquivo (str): Caminho completo do arquivo Excel.
        aba (str): Nome da aba onde salvar os dados.

    Retorno:
        str: Mensagem de sucesso.
        Exception: PermissionError caso o arquivo esteja aberto.
    """
    # garante que a pasta existe
    os.makedirs(os.path.dirname(arquivo), exist_ok=True)

    # abre ou cria o arquivo Excel
    if os.path.exists(arquivo):
        wb = load_workbook(arquivo)
        if aba in wb.sheetnames:
            ws = wb[aba]
        else:
            ws = wb.create_sheet(aba)
            ws.append(["Hora Inicial", "Hora Final", "Total Execução (min)", "Robô"])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = aba
        ws.append(["Hora Inicial", "Hora Final", "Total Execução (min)", "Robô"])

    # converte timestamps em horas legíveis
    hora_inicial = datetime.fromtimestamp(start_time).strftime("%Y-%m-%d %H:%M:%S")
    hora_final = datetime.fromtimestamp(end_time).strftime("%Y-%m-%d %H:%M:%S")
    tempo_total_min = round((end_time - start_time) / 60, 2)
    tempo_total_min = f'00:{tempo_total_min}'.replace('.', ':')

    # grava uma linha
    ws.append([hora_inicial, hora_final, tempo_total_min, nome_robo])

    try:
        wb.save(arquivo)
        return "Indicador salvo com sucesso!"
    except PermissionError:
        raise PermissionError(f"Feche o arquivo '{arquivo}' no Excel e tente novamente.")


def espera_personalizada(*acoes, retorno=False, inicio_random=3, fim_random=8):
    """
    Executa uma ou mais ações com tentativas repetidas até que tenham sucesso. Caso tenha retorno, adicione a função que deve receber retorno como a última.

    Parâmetros:
        *acoes (callable): Funções ou lambdas que executam ações.
        **retorno (bool): Variável boolena para verificar se é necessário retornar algo.

    Retorno:
        None
    """
    contador = 0
    while True:
        try:
            # espera aleatória para simular comportamento humano
            sleep(random.randint(inicio_random, fim_random))

            # executa todas menos a última
            for acao in acoes[:-1] if retorno else acoes:
                acao()

            # se tiver retorno, executa a última e devolve
            if retorno:
                return acoes[-1]()
            
            break
        except Exception:
            if contador == 2:
                break
            contador += 1
            continue


def sair_ihs(driver):
    driver.switch_to.default_content()
    driver.find_element(By.CSS_SELECTOR, 'div.pull-right div.logout a').click()

def clica_em_prosseguir(driver):
    driver.find_element(By.CSS_SELECTOR, 'div.div-token center button.btn-default.btn.button-token').click()

def pega_texto_elemento(driver, primeira_parte_id, iterador):
    id_elemento = ajusta_id(primeira_parte_id, iterador)
    elemento = espera_personalizada(
        lambda: driver.find_element(By.ID, id_elemento),
        retorno=True,
        inicio_random=1,
        fim_random=3
    )

    return elemento

def ler_valores_extrato(user, ontem, parse_valor):
    pasta = r'\\172.17.67.14\findev$\Automação - CDC\\'

    # Mapeamento por loja
    config = {
        'JUAZEIRO': {
            'nome': f'JUA ITAU {ontem.strftime("%d")}.xlsx',
            'header': 9,
            'col_razao': 'Razão Social',
            'col_valor': 'Valor (R$)',
            'filtro_texto': 'BANCO HONDA'
        },
        'TERRA_SANTA': {
            'nome': f'INHA ITAU {ontem.strftime("%d")}.xlsx',  # confirme se é "INHA" mesmo
            'header': 9,
            'col_razao': 'Razão Social',
            'col_valor': 'Valor (R$)',
            'filtro_texto': 'BANCO HONDA'
        },
        'NOVA_ONDA': {
            'nome': f'NO ITAU {ontem.strftime("%d")}.xlsx',
            'header': 9,
            'col_razao': 'Razão Social',
            'col_valor': 'Valor (R$)',
            'filtro_texto': 'BANCO HONDA'
        },
        'CRATO': {
            'nome': f'JUA ARBI {ontem.strftime("%d")}.xlsx',
            'header': 8,
            'col_razao': 'Nome Contraparte',
            'col_valor': 'Valor',
            'filtro_texto': 'BANCO HONDA'
        },
    }

    if user.nome_loja not in config:
        raise ValueError(f"Loja '{user.nome_loja}' não mapeada.")

    cfg = config[user.nome_loja]
    nome_esperado = cfg['nome']

    try:
        arquivos = os.listdir(pasta)
    except Exception as e:
        messagebox.showerror("Erro", f"Não foi possível listar a pasta.\n{e}")
        return None

    # 1) Procura somente UMA vez pelo arquivo esperado
    caminho_arquivo = None
    for arquivo in arquivos:
        if arquivo == nome_esperado:
            caminho_arquivo = os.path.join(pasta, arquivo)
            break

    # 2) Se não encontrou, pergunta uma vez e re-checa
    while caminho_arquivo is None:
        resp = messagebox.askokcancel(
            "Arquivo não encontrado",
            f"Não existe o arquivo da loja {user.nome_loja} na pasta.\n"
            f"Esperado: {nome_esperado}\n\nClique OK após adicionar o arquivo, ou Cancelar para sair."
        )
        if not resp:
            return None  # ou lance exceção / retorne indicador para sair

        # Recarrega a lista e tenta novamente
        try:
            arquivos = os.listdir(pasta)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível listar a pasta.\n{e}")
            return None

        for arquivo in arquivos:
            if arquivo == nome_esperado:
                caminho_arquivo = os.path.join(pasta, arquivo)
                break

        # Se ainda não achou, o while repete e pergunta de novo

    # 3) Lê a planilha e filtra
    try:
        df = pandas.read_excel(caminho_arquivo, header=cfg['header'])
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler o Excel '{nome_esperado}'.\n{e}")
        return None

    col_razao = cfg['col_razao']
    col_valor = cfg['col_valor']
    filtro_txt = cfg['filtro_texto']

    if col_razao not in df.columns or col_valor not in df.columns:
        messagebox.showerror(
            "Colunas não encontradas",
            f"As colunas esperadas não foram encontradas no arquivo:\n"
            f"- {col_razao}\n- {col_valor}\n\nVerifique o header (linha de título) configurado."
        )
        return None

    filtro = df[col_razao].astype(str).str.contains(filtro_txt, case=False, na=False)

    # 4) Converte os valores para float com parse_valor SEMPRE
    try:
        if user.nome_loja == 'CRATO':
            valores_extrato = df.loc[filtro, col_valor].apply(parse_valor).tolist()
        else:
            valores_extrato = df.loc[filtro, col_valor].tolist()
    except Exception as e:
        messagebox.showerror(
            "Erro de conversão",
            f"Falha ao converter os valores da coluna '{col_valor}'.\n{e}"
        )
        return None

    return valores_extrato

load_dotenv()

def get_all_users(lojas: str):
    """
    Carrega todos os usuários definidos no arquivo `.env` e cria objetos `User`.

    Parâmetros:
        None

    Retorno:
        list[User]: Lista de instâncias de usuários com os atributos
        `codigo`, `usuario`, `senha` e `nome_loja`.
    """
    usuarios = []
    prefixos = tela(lojas)

    for store, code, user, password in zip(prefixos['LOJAS'], prefixos['CODIGOS'], prefixos['USUARIOS'], prefixos['SENHAS']):
        usuarios.append(User(codigo=code, usuario=user, senha=password, nome_loja=store))
    return usuarios

load_dotenv()

# 1) Constante de lojas (fonte da verdade)
LOJAS: List[str] = [
    "JUAZEIRO",
    "TERRA_SANTA",
    "NOVA_ONDA",
    "CRATO",
]

# 2) Normalização -> "Terra Santa" / "terra-santa" / "terra_santa" => "TERRA_SANTA"
def _norm_key(s: str) -> str:
    if s is None:
        return ""
    s = s.strip()
    # remove acentos
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = s.upper()
    # troca qualquer coisa não alfanumérica por "_"
    s = re.sub(r"[^A-Z0-9]+", "_", s)
    # remove underscores duplicados e bordas
    s = re.sub(r"_+", "_", s).strip("_")
    return s

# 3) Mapa de normalizado -> nome canônico da constante LOJAS
_LOJAS_MAP = { _norm_key(nome): nome for nome in LOJAS }

def _parse_param_lojas(lojas_param: str) -> List[str]:
    """
    Converte o parâmetro 'lojas' em lista de nomes CANÔNICOS (como em LOJAS),
    preservando a ordem de entrada e removendo duplicatas pela 1ª ocorrência.
    """
    if not lojas_param:
        return []

    if lojas_param.lower() == "all":
        # Mantém a ordem definida na constante LOJAS
        return list(LOJAS)

    vistos = set()
    ordem_canonica: List[str] = []

    for token in lojas_param.split(","):
        key = _norm_key(token)
        if not key:
            continue
        if key not in _LOJAS_MAP:
            # coletamos invalida e tratamos depois
            ordem_canonica.append(f"!INVALIDA:{key}")
        else:
            can = _LOJAS_MAP[key]
            if can not in vistos:
                ordem_canonica.append(can)
                vistos.add(can)

    return ordem_canonica

def _env_or_fail(varname: str) -> str:
    val = os.getenv(varname)
    if val is None or str(val).strip() == "":
        raise ValueError(f"Variável de ambiente ausente/vázia: {varname}")
    return str(val).strip()

def tela(lojas: str) -> Dict[str, List[str]]:
    """
    Retorna dicionário com listas alinhadas:
      {'LOJAS': [...], 'CODIGOS': [...], 'USUARIOS': [...], 'SENHAS': [...]}

    Parâmetro:
      lojas (str): 'all' ou valores separados por vírgula. Aceita variações
                   ('Terra Santa', 'terra-santa', 'TERRA_SANTA' etc.).

    Regras:
      - Usa a constante LOJAS como verdade absoluta dos nomes canônicos.
      - Busca credenciais do .env nas chaves:
          CODIGO_<KEY>, USUARIO_<KEY>, SENHA_<KEY>
        onde <KEY> é a forma normalizada: p.ex., 'TERRA_SANTA'.
      - Preserva ordem de entrada e ignora duplicatas por 1ª ocorrência.
      - Erra com mensagem clara se loja inválida ou env ausente.
    """
    ordem = _parse_param_lojas(lojas)

    if not ordem:
        raise ValueError("Nenhuma loja informada/derivada do parâmetro.")

    # Validação de lojas inválidas
    lojas_invalidas = [x.split(":", 1)[1] for x in ordem if x.startswith("!INVALIDA:")]
    if lojas_invalidas:
        # traduz a key inválida para formato legível
        legiveis = [x.replace("_", " ").title() for x in lojas_invalidas]
        raise ValueError(
            f"Lojas inválidas (não constam em LOJAS): {legiveis}. "
            f"Válidas: {', '.join(LOJAS)}"
        )

    # Remover quaisquer marcadores inválidos (se chegasse até aqui)
    ordem = [x for x in ordem if not x.startswith("!INVALIDA:")]

    lojas_out: List[str] = []
    codigos: List[str] = []
    usuarios: List[str] = []
    senhas: List[str] = []

    for loja_canon in ordem:
        key = _norm_key(loja_canon)  # ex.: "TERRA_SANTA"
        try:
            codigo  = _env_or_fail(f"CODIGO_{key}")
            usuario = _env_or_fail(f"USUARIO_{key}")
            senha   = _env_or_fail(f"SENHA_{key}")
        except ValueError as ve:
            # acrescenta contexto de qual loja falhou
            raise ValueError(f"[{loja_canon}] {ve}") from ve

        lojas_out.append(loja_canon)  # mantém o canônico da constante
        codigos.append(codigo)
        usuarios.append(usuario)
        senhas.append(senha)

    return {
        "LOJAS": lojas_out,
        "CODIGOS": codigos,
        "USUARIOS": usuarios,
        "SENHAS": senhas,
    }


def conciliacao_cdc_honda_main(lojas: str):
        # Busca todos os usuários cadastrados
    try:
        usuarios = get_all_users(lojas)
    except Exception as e:
        return False, f'Erro ao buscar os usuários.\nDescrição: {str(e)}'
    # Marca o início da execução (para métricas)
    start_time = time()

    # Configura o WebDriver
    driver, wdw, PASTA_DOWNLOADS = _ensure_driver()

    url = 'https://www3.honda.com.br/corp/ihs/portal/#/login'
    driver.get(url)  # Entra no site do IHS

    for user in usuarios:
        try:


            driver.delete_all_cookies()

            # --- Login ---
            espera_personalizada(inicio_random=1, fim_random=3)
            driver.find_element(By.CSS_SELECTOR, '#codEmpresa').send_keys(user.codigo)
            espera_personalizada(inicio_random=1, fim_random=3)
            driver.find_element(By.CSS_SELECTOR, '#codUsuario').send_keys(user.usuario)
            espera_personalizada(inicio_random=1, fim_random=3)
            driver.find_element(By.CSS_SELECTOR, '#senha').send_keys(user.senha)
            espera_personalizada(inicio_random=1, fim_random=3)

            submit = wdw.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '#submitLogin'))
            )
            submit.click()
            
            prosseguir = wdw.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'div.div-token center button.btn-default.btn.button-token'))
            )

            prosseguir.click()

            # --- Navegação pelo menu ---
            espera_personalizada()

            try:
                btn_mensagem = WebDriverWait(driver, timeout=15).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, 'div#container-portal button[type="submit"]'))
                )

                btn_mensagem.click()
            except:
                pass
            
            clica_na_aba(wdw, '.a-empresaHonda', 'banco')  # Financiamento
            espera_personalizada(inicio_random=1, fim_random=3)
            clica_na_aba(wdw, '.menugeral ul li a', 'gestão de financiamentos')

            espera_personalizada(inicio_random=1, fim_random=3)
            clica_na_aba(wdw, '#sideMenu0 .itemMenu a', 'gestão de financiamentos')
            espera_personalizada(inicio_random=1, fim_random=3)
            clica_na_aba(wdw, '#sideMenu0 .itemMenu a', 'controlar pagamento')
            espera_personalizada(inicio_random=1, fim_random=3)
            clica_na_aba(wdw, '#sideMenu0 .itemMenu a', 'consultar cc concessionaria')

            # --- Acessa o frame principal ---
            espera_personalizada(
                lambda: driver.switch_to.default_content(),
                lambda: driver.switch_to.frame(0),
                lambda: driver.find_element(By.CSS_SELECTOR, 'select').click(),
                inicio_random=2,
                fim_random=4
            )

            espera_personalizada(inicio_random=1, fim_random=3)
            clica_na_aba(wdw, 'select option', 'financiamento de bens')

            # --- Define data de ontem ---
            ontem = datetime.now() - timedelta(days=1)
            if ontem.weekday() == 6:  # se for domingo, volta 2 dias (sexta-feira)
                ontem = ontem - timedelta(days=2)
            data_ontem = ontem.strftime("%d/%m/%y")

            sleep(random.randint(2, 4))

            # --- Preenche período da consulta ---
            preencher_campo(driver, (By.CSS_SELECTOR, '#vWDPERINI'), data_ontem)
            preencher_campo(driver, (By.CSS_SELECTOR, '#vWDPERFIM'), data_ontem)

            # Confirma pesquisa
            clicar_pelo_atributo(
                driver=driver,
                atributo='value',
                texto_comparar='confirmar',
                locator=(By.CSS_SELECTOR, 'input')
            )

            # --- Busca dados na Honda ---
            dados = {}
            lista_documentos_nao_pagos = []
            logs = []
            documentos = []

            # ===================================
            #    Pega os valores do extrato
            # ===================================
            espera_personalizada(lambda: driver.execute_script("document.body.style.zoom='57%'"), inicio_random=2,fim_random=4)

            try:

                valores_extrato = ler_valores_extrato(user, ontem, parse_valor)

            except Exception as e:
                logs.append(f'🚫 Erro ao ler o arquivo do extrato. Ele não existe ou não é um arquivo Excel.\nDescrição: {str(e)}')

                sair_ihs(driver)
                continue

            
            # ===================================
            #   Pega os dados se for uma página
            # ===================================
            espera_personalizada(inicio_random=2, fim_random=5)
            lote_teste = pega_texto_elemento(driver, 'span_vCCNNLOT', 15)
            if lote_teste == None:
                for i in range(1,15):
                    if i == 1: espera_personalizada(inicio_random=4, fim_random=8)
                    
                    try:
                        documento = pega_texto_elemento(driver, 'span_vCCNNDOC', i)

                        if documento == None: break

                        documentos.append(documento.text.strip())
                    except:
                        continue
                
                contagem = Counter(documentos)
                documentos_nao_pagos = [item for item, qtd in contagem.items() if qtd == 1]

                if documentos_nao_pagos:
                    logs.append(f'{'=' * 60}\n')
                    for doc in documentos_nao_pagos:
                        for i in range(1,15):
                            if i == 1: espera_personalizada(inicio_random=4, fim_random=8)

                            documento = pega_texto_elemento(driver, 'span_vCCNNDOC', i)

                            if documento == None: break

                            if doc == documento.text.strip():
                                lote = pega_texto_elemento(driver, 'span_vCCNNLOT', i)

                                if lote.text.strip() != '0':
                                    valor = pega_texto_elemento(driver, 'span_vCCNVEVT', i)

                                    valor_parseado = parse_valor(valor.text.strip())

                                    if valor_parseado not in valores_extrato:
                                        logs.append(f'❌ Documento não pago: {doc}')

                    logs.append(f'{'=' * 60}\n')
                
                for i in range(1,15):
                    if i == 1: espera_personalizada(inicio_random=4, fim_random=8)

                    lote = pega_texto_elemento(driver, 'span_vCCNNLOT', i)

                    if lote == None: break

                    if lote.text.strip() != '0':
                        valor = pega_texto_elemento(driver, 'span_vCCNVEVT', i)

                        valor_parseado = parse_valor(valor.text.strip())

                        if valor_parseado in valores_extrato:
                            lote.click()

                            nomes, valores = espera_personalizada(
                                lambda: pega_valores_liquidos(driver, wdw, EC),
                                retorno=True
                            )

                            if nomes and valores:
                                for nome, valor in zip(nomes, valores):
                                    dados[nome] = parse_valor(valor)
            else:
            # ===================================
            #   Pega os dados se for duas página
            # ===================================
                for p in range(2):
                    for i in range(1,16):
                        if i == 1: espera_personalizada(inicio_random=4, fim_random=8)

                        try:
                            documento = pega_texto_elemento(driver, 'span_vCCNNDOC', i)

                            if documento == None: break

                            documentos.append(documento.text.strip())
                        except:
                            continue
                
                    if p == 0:
                        tenta_passar_pagina(driver)
                        espera_personalizada(inicio_random=4, fim_random=8)
                    else:
                        tenta_passar_pagina_atras(driver)
                        espera_personalizada(inicio_random=4, fim_random=8)

                contagem = Counter(documentos)
                documentos_nao_pagos = [item for item, qtd in contagem.items() if qtd == 1]

                quantas_pular = 2
                if lote_teste == '0':
                    quantas_pular = 1

                for p in range(quantas_pular):
                    if documentos_nao_pagos:
                        logs.append(f'{'=' * 60}\n')
                        for doc in documentos_nao_pagos:
                            for i in range(1,16):
                                if i == 1: espera_personalizada(inicio_random=4, fim_random=8)

                                documento = pega_texto_elemento(driver, 'span_vCCNNDOC', i)

                                if documento == None: break

                                if doc == documento.text.strip():
                                    lote = pega_texto_elemento(driver, 'span_vCCNNLOT', i)

                                    if lote.text.strip() != '0':
                                        valor = pega_texto_elemento(driver, 'span_vCCNVEVT', i)

                                        valor_parseado = parse_valor(valor.text.strip())

                                        if valor_parseado not in valores_extrato:
                                            logs.append(f'❌ Documento não pago: {doc}')

                        logs.append(f'{'=' * 60}\n')
                    
                    for i in range(1,16):
                        if i == 1: espera_personalizada(inicio_random=4, fim_random=8)

                        lote = pega_texto_elemento(driver, 'span_vCCNNLOT', i)

                        if lote == None: break

                        if lote.text.strip() != '0':
                            valor = pega_texto_elemento(driver, 'span_vCCNVEVT', i)

                            valor_parseado = parse_valor(valor.text.strip())

                            if valor_parseado in valores_extrato:
                                lote.click()

                                nomes, valores = espera_personalizada(
                                    lambda: pega_valores_liquidos(driver, wdw, EC),
                                    retorno=True
                                )

                                if nomes and valores:
                                    for nome, valor in zip(nomes, valores):
                                        dados[nome] = parse_valor(valor)
                    
                    if quantas_pular == 2:
                        if p == 0:
                            tenta_passar_pagina(driver)
                            espera_personalizada(inicio_random=4, fim_random=8)
                        else:
                            tenta_passar_pagina_atras(driver)
                            espera_personalizada(inicio_random=4, fim_random=8)

        except Exception as e:
            logs.append(f'🚫 Erro ao buscar os dados na loja de {user.nome_loja}.\nDescrição: {str(e)}')
            
            sair_ihs(driver)
            continue

        try:
            # --- Busca dados no banco local ---
            cur = busca_dados_db()

            
            titulos, duplicatas, valores, nomes, situacoes = [], [], [], [], []

            for row in cur:
                try:
                    eh_compativel, log, valor_normalizado = compara_valores(
                        nome_db=unidecode(row[-1].strip().upper()),
                        valor_db=row[3],
                        empresa=row[-2],
                        titulo=row[1],
                        dados_honda=dados
                    )

                    if not log or '❌ Nenhum cliente correspondente encontrado.' in log: continue

                    logs.append(log)

                    # Se for compatível, guarda para exportação
                    if eh_compativel:
                        valor_ajustado_real = f'{valor_normalizado:,.2f}'.replace('.', '_').replace(',', '.').replace('_', ',')
                        titulos.append(row[1])
                        duplicatas.append(row[2])
                        valores.append(valor_ajustado_real)
                        nomes.append(row[-1])
                        if '❌' in log:
                            situacoes.append('Incompatível por causa do valor')
                        else:
                            situacoes.append('Compatível')
                except TypeError:
                    continue
        except Exception as e:
            logs.append(f'🚫 Erro ao consultar dados no banco de dados.\nDescrição: {str(e)}')
            
            sair_ihs(driver)
            continue

        try:
            # --- Monta DataFrame com os dados compatíveis ---
            cabecalhos = ['TITULO', 'DUPLICATA', 'VALOR']
            cabecalhos_excel = ['NOME', 'TITULO', 'DUPLICATA', 'VALOR', 'SITUACAO']
            data = {
                cabecalhos[0]: titulos,
                cabecalhos[1]: duplicatas,
                cabecalhos[2]: valores
            }
            data_excel = {
                cabecalhos_excel[0]: nomes,
                cabecalhos_excel[1]: titulos,
                cabecalhos_excel[2]: duplicatas,
                cabecalhos_excel[3]: valores,
                cabecalhos_excel[4]: situacoes,
            }
            df = pandas.DataFrame(data, columns=cabecalhos)
            df_excel = pandas.DataFrame(data_excel, columns=cabecalhos_excel)

            hoje = datetime.now().strftime("%d-%m-%Y")

            # Salva logs e CSV por loja
            cria_arquivo_log(user.nome_loja, logs)
            # Caminhos base usando Path (melhor que strings com \\)
            base_dir = P(r"\\172.17.67.14\findev$\Automação - CDC")
            csv_dir = base_dir / "Arquivos_csv"
            excel_dir = base_dir / "Arquivos_excel"

            # Monta os nomes de arquivos de forma segura e legível
            arquivo_csv = csv_dir / f"consolidado_{user.nome_loja}_{hoje}.csv"
            arquivo_excel = excel_dir / f"resumo_{user.nome_loja}_{hoje}.xlsx"

            # Salva os arquivos
            df.to_csv(arquivo_csv, index=False, sep=';', encoding='utf-8-sig')
            df_excel.to_excel(arquivo_excel, index=False)

            sair_ihs(driver)

            espera_personalizada(inicio_random=2, fim_random=4)
        except Exception as e:
            logs.append(f'🚫 Erro ao gerar o arquivo .csv da loja {user.nome_loja}.\nDescrição: {str(e)}')
            
            sair_ihs(driver)
            continue

    # --- Finalização ---
    end_time = time()
    tempo_total_min = str(timedelta(seconds=end_time - start_time))

    print(f'Progama finalizado em {tempo_total_min}m')

    try:
        resposta = cria_indicador_de_tempo_execucao(start_time, end_time, 'CONCILIAÇÃO DB COM HONDA')
        print(resposta)
    except Exception as e:
        return False, f'Erro ao registrar indicador de tempo de execução.\nDescrição: {str(e)}'

    return True, 'Conciliação do CDC Hond realizado corretamente.'