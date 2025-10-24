from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook, Workbook
from selenium.webdriver.common.by import By
from APP.DTO.ihs_dto import User
from tkinter import messagebox
from datetime import datetime
from typing import Dict, List
from pathlib import Path as P
from pathlib import PurePath
from time import sleep
import pandas as pd
import pdfplumber
import zipfile
import random
import shutil
import os
from selenium.webdriver.common.by import By
from time import sleep, time
import random
from APP.Config.ihs_config import _ensure_driver
from APP.Core.baixa_arquivos_core import Path
from selenium.webdriver.support import expected_conditions as EC
import logging
from datetime import datetime, timedelta
from selenium.webdriver.support.ui import WebDriverWait

ARQ = r'\\172.17.67.14\findev$\Automação - CNH\SENHAS IHS.xlsx'

def preencher_campo(driver, value, texto):
    """
    Preenche um campo de texto em um formulário web.

    Parâmetros:
        driver (webdriver): Instância do Selenium WebDriver.
        value (tuple): Localizador do elemento (By, valor).
        texto (str): Texto a ser inserido no campo.

    Retorno:
        None
    """
    campo = driver.find_element(By.CSS_SELECTOR, value)
    campo.click()
    espera_personalizada(
        lambda: campo.send_keys(texto),
        inicio_random=2,
        fim_random=4
    )


def clicar_em_um_botao(driver, value):
    """
    Clica em um botão simulando a tecla ENTER.

    Parâmetros:
        driver (webdriver): Instância do Selenium WebDriver.
        value (tuple): Localizador do botão (By, valor).

    Retorno:
        None
    """
    espera_personalizada(
        lambda: driver.find_element(By.CSS_SELECTOR, value).send_keys(Keys.ENTER),
        inicio_random=2,
        fim_random=4
    )


def clica_na_aba(wdw, css_selector, texto_ancora):
    """
    Clica em uma aba específica com base no texto exibido.

    Parâmetros:
        driver (webdriver): Instância do Selenium WebDriver.
        css_selector (str): Seletor CSS para localizar as abas.
        texto_ancora (str): Texto da aba desejada (case-insensitive).

    Retorno:
        None
    """
    aba = wdw.until(
        lambda d: next(
            (
                el for el in d.find_elements(By.CSS_SELECTOR, css_selector)
                if texto_ancora in el.text.strip().lower()
            ),
            None
        )
    )

    if aba:
        aba.click()
    else:
        print("❌ Aba com texto não encontrada.")

def clica_em_um_botao_pelo_nome(driver, value, texto_ancora):
    """
    Clica em uma aba específica com base no texto exibido.

    Parâmetros:
        driver (webdriver): Instância do Selenium WebDriver.
        css_selector (str): Seletor CSS para localizar as abas.
        texto_ancora (str): Texto da aba desejada (case-insensitive).

    Retorno:
        None
    """
    elementos = driver.find_elements(By.CSS_SELECTOR, value)
    for a in elementos:
        if a.text.strip().lower() == texto_ancora:
            espera_personalizada(
                lambda: a.click()
            )
            break


def clicar_pelo_atributo(driver, atributo, texto_comparar, value):
    """
    Clica em um elemento com base no valor de um atributo.

    Parâmetros:
        driver (webdriver): Instância do Selenium WebDriver.
        atributo (str): Nome do atributo HTML a ser comparado.
        texto_comparar (str): Valor esperado do atributo.
        value (tuple): Localizador do grupo de elementos (By, valor).

    Retorno:
        None
    """
    entradas = driver.find_elements(By.CSS_SELECTOR, value)
    for i in entradas:
        if i.get_attribute(atributo).lower() == texto_comparar:
            espera_personalizada(
                lambda: i.click(),
                inicio_random=1,
                fim_random=3
            )
            break



def cria_indicador_de_tempo_execucao(
    start_time: float,
    end_time: float,
    nome_robo: str,
    arquivo: str = r"\\172.17.67.14\findev$\indicadores.xlsx",
    aba: str = "Execuções",
):
    """
    Cria ou atualiza um arquivo Excel com indicadores de tempo de execução.

    Parâmetros:
        start_time (float): Timestamp do início.
        end_time (float): Timestamp do fim.
        nome_robo (str): Nome do robô que executou.
        arquivo (str): Caminho completo do arquivo Excel.
        aba (str): Nome da aba onde salvar os dados.

    Retorno:
        str: Mensagem de sucesso.
        Exception: PermissionError caso o arquivo esteja aberto.
    """
    # garante que a pasta existe
    os.makedirs(os.path.dirname(arquivo), exist_ok=True)

    # abre ou cria o arquivo Excel
    if os.path.exists(arquivo):
        wb = load_workbook(arquivo)
        if aba in wb.sheetnames:
            ws = wb[aba]
        else:
            ws = wb.create_sheet(aba)
            ws.append(["Hora Inicial", "Hora Final", "Total Execução (min)", "Robô"])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = aba
        ws.append(["Hora Inicial", "Hora Final", "Total Execução (min)", "Robô"])

    # converte timestamps em horas legíveis
    hora_inicial = datetime.fromtimestamp(start_time).strftime("%Y-%m-%d %H:%M:%S")
    hora_final = datetime.fromtimestamp(end_time).strftime("%Y-%m-%d %H:%M:%S")
    tempo_total_min = round((end_time - start_time) / 60, 2)
    tempo_total_min = f'00:{tempo_total_min}'.replace('.', ':')

    # grava uma linha
    ws.append([hora_inicial, hora_final, tempo_total_min, nome_robo])

    try:
        wb.save(arquivo)
        return "Indicador salvo com sucesso!"
    except PermissionError:
        raise PermissionError(f"Feche o arquivo '{arquivo}' no Excel e tente novamente.")


def espera_personalizada(*acoes, retorno=False, inicio_random=3, fim_random=8):
    """
    Executa uma ou mais ações com tentativas repetidas até que tenham sucesso. Caso tenha retorno, adicione a função que deve receber retorno como a última.

    Parâmetros:
        *acoes (callable): Funções ou lambdas que executam ações.
        **retorno (bool): Variável boolena para verificar se é necessário retornar algo.

    Retorno:
        None
    """
    contador = 0
    while True:
        try:
            # espera aleatória para simular comportamento humano
            sleep(random.randint(inicio_random, fim_random))

            # executa todas menos a última
            for acao in acoes[:-1] if retorno else acoes:
                acao()

            # se tiver retorno, executa a última e devolve
            if retorno:
                return acoes[-1]()
            
            break
        except Exception:
            if contador == 5:
                break
            contador += 1
            continue


def sair_ihs(driver, value):
    driver.switch_to.default_content()
    driver.find_element(By.CSS_SELECTOR, value).click()

def clica_em_prosseguir(driver, value):
    driver.find_element(By.CSS_SELECTOR, value).click()

def ajusta_id(primeira_parte, numero_lote):
    """
    Ajusta o ID de um elemento conforme o número do lote.

    Parâmetros:
        primeira_parte (str): Prefixo do ID.
        numero_lote (int): Número do lote.

    Retorno:
        str: ID ajustado conforme o tamanho do número.
    """
    if numero_lote < 10:
        return f'{primeira_parte}_000{numero_lote}'
    elif numero_lote > 9 and numero_lote < 100:
        return f'{primeira_parte}_00{numero_lote}'
    elif numero_lote > 99 and numero_lote < 1000:
        return f'{primeira_parte}_0{numero_lote}'
    elif numero_lote > 999 and numero_lote < 10000:
        return f'{primeira_parte}_{numero_lote}'


def pega_texto_elemento(driver, primeira_parte_id, iterador):
    id_elemento = ajusta_id(primeira_parte_id, iterador)
    elemento = espera_personalizada(
        lambda: driver.find_element(By.CSS_SELECTOR, id_elemento),
        retorno=True
    )

    return (id_elemento, elemento.text)

def troca_nome_arquivo(PASTA_DOWNLOADS, nome_arquivo_destino):
    espera_personalizada(inicio_random=8, fim_random=12)
    lista_arquivos_baixados = sorted(
        [
            f for f in os.listdir(PASTA_DOWNLOADS)
            if os.path.isfile(os.path.join(PASTA_DOWNLOADS, f))
        ],
        key=lambda f: os.path.getmtime(os.path.join(PASTA_DOWNLOADS, f))
    )

    if lista_arquivos_baixados:
        origem = PASTA_DOWNLOADS / lista_arquivos_baixados[-1]
        destino = PASTA_DOWNLOADS / nome_arquivo_destino
        origem.replace(destino)


def seleciona_uma_aba_do_navegador(driver, numero_janela, eh_para_fechar=False):
    abas = espera_personalizada(
        lambda: driver.window_handles,
        retorno=True
    )
    
    driver.switch_to.window(abas[numero_janela])
    if not eh_para_fechar:
        driver.set_window_position(-3566,-3566)
        driver.set_window_size(1366, 768)

def extract_text_pdfplumber(nome_loja, pdf_path=r'\\172.17.67.14\findev$\Automação - CNH\Baixa de Arquivos\Arquivos Baixados'):
    paginas = []
    pdf_path = P(pdf_path)
    pdf_file_path = pdf_path.joinpath(f'{nome_loja}.pdf')
    with pdfplumber.open(pdf_file_path) as pdf:
        for page in pdf.pages:
            paginas.append(page.extract_text() or "")

    numero_nota = []
    valor = []
    grupo_cota_rd = []

    for i in range(len(paginas)):
        lista_linhas = paginas[i].split('\n')
        for j in lista_linhas:
            if 'LANC. PAGAMEN. AUTOM' in j:
                texto_lista = str(j).split(sep=' ')
                numero_nota.append(texto_lista[1])
                valor.append(texto_lista[-2])

                grupo = texto_lista[2].split(sep='/')
                if len(grupo) < 4:
                    grupo_cota_rd.append(texto_lista[2] + texto_lista[3].strip())
                else:
                    grupo_cota_rd.append(texto_lista[2])

    data = {
        'NUMERO NOTA': numero_nota,
        'VALOR': valor,
        'GRUPO COTA RD': grupo_cota_rd
    }

    df = pd.DataFrame(data, columns=['NUMERO NOTA', 'VALOR', 'GRUPO COTA RD'])

    caminho_excel = pdf_path.joinpath(f'extracao_pdf_{nome_loja}.xlsx')

    df.to_excel(caminho_excel, index=False)

def espera_mensagem(driver, value, wdw, EC, timeout=180):
        # Espera até que o elemento esteja visível
        while True:
            try:
                elemento = wdw(driver, timeout).until(
                    EC.invisibility_of_element((By.CSS_SELECTOR, value))
                )
                break
            except:
                continue

def extrair_arquivo_alvo(zip_path, destino, nome_arquivo=None, prefixo=None):
    """
    - nome_arquivo: nome exato do arquivo (ex.: 'meu_arquivo.txt') — opcional
    - prefixo: caminho dentro do ZIP onde procurar (ex.: 'hda0334_new/d$/wwwhondaihs/internet/dwnhsfzip/') — opcional
    Extrai o arquivo para 'destino' SEM subpastas (flatten).
    """
    destino = P(destino)
    destino.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(zip_path) as z:
        nomes = [n.replace('\\', '/') for n in z.namelist() if not n.endswith('/')]

        # Filtra pelo prefixo se informado
        if prefixo:
            prefixo = prefixo.replace('\\', '/').rstrip('/') + '/'
            nomes = [n for n in nomes if n.startswith(prefixo)]

        if nome_arquivo:
            # Busca por nome exato no final do caminho
            candidatos = [n for n in nomes if PurePath(n).name.lower() == nome_arquivo.lower()]
        else:
            # Sem nome específico: pega o arquivo mais "profundo"
            candidatos = sorted(nomes, key=lambda p: p.count('/'), reverse=True)

        if not candidatos:
            raise FileNotFoundError("Arquivo não encontrado dentro do ZIP com os filtros dados.")

        escolhido = candidatos[0]  # ou aplique outra regra (maior, mais novo, etc.)
        out_path = destino / PurePath(escolhido).name

        with z.open(escolhido) as src, open(out_path, 'wb') as dst:
            shutil.copyfileobj(src, dst)

        return out_path
    
def get_all_users(lojas):
    """
    Carrega todos os usuários definidos no arquivo `.env` e cria objetos `User`.

    Parâmetros:
        None

    Retorno:
        list[User]: Lista de instâncias de usuários com os atributos
        `codigo`, `usuario`, `senha` e `nome_loja`.
    """
    usuarios = []
    prefixos = tela(lojas)

    # Validação simples
    tamanhos = {k: len(v) for k, v in prefixos.items()}
    if len(set(tamanhos.values())) != 1:
        raise ValueError(f"Listas com tamanhos diferentes: {tamanhos}")

    for l, c, u, s in zip(prefixos['LOJAS'], prefixos['CODIGOS'], prefixos['USUARIOS'], prefixos['SENHAS']):
        codigo = c
        usuario = u
        senha = s
        nome_loja = l

        if codigo and usuario and senha:
            usuarios.append(User(codigo=codigo, usuario=usuario, senha=senha, nome_loja=nome_loja))
    return usuarios

def _norm_loja(nome: str) -> str:
    return str(nome).strip().upper().replace(" ", "-")

def _normaliza_lista_lojas_param(param: str) -> List[str]:
    """
    Converte o parâmetro 'lojas' em uma lista de keys normalizadas.
    Ex.: 'Ares Motos, Cajazeiras' -> ['ares-motos', 'cajazeiras'] (exemplo)
    """
    partes = [p.strip() for p in param.split(",") if p.strip()]
    return [_norm_loja(p) for p in partes]

def tela(lojas: str) -> Dict[str, List[str]]:
    """
    Lê o Excel e retorna o dicionário de listas filtrado:
    {'LOJAS': [...], 'CODIGOS': [...], 'USUARIOS': [...], 'SENHAS': [...]}

    Parâmetros:
        lojas (str): 'all' ou lista separada por vírgula (ordem preservada).
                     Aceita tanto nome exibido quanto key normalizada.

    Retorno:
        dict[str, list[str]]
    """
    # 1) Lê o Excel
    df = pd.read_excel(ARQ, dtype=str)  # dtype=str evita NaN -> float
    esperadas = {"LOJAS", "CODIGOS", "USUARIOS", "SENHAS"}
    if not esperadas.issubset(set(df.columns)):
        raise ValueError(f"Colunas esperadas: {esperadas} | Encontradas: {set(df.columns)}")

    # 2) Cria a coluna de chave normalizada
    df["_LOJA_KEY"] = df["LOJAS"].map(_norm_loja)

    # 3) Determina quais keys usar e a ordem
    if lojas.lower() == "all":
        # usa TODAS as lojas na ordem do arquivo (removendo duplicatas por primeira ocorrência)
        # se houver duplicatas, consideramos a primeira ocorrência
        ordem_keys = []
        vistos = set()
        for k in df["_LOJA_KEY"]:
            if k not in vistos:
                ordem_keys.append(k)
                vistos.add(k)
    else:
        # usa SOMENTE as lojas passadas na URL, preservando a ordem fornecida
        ordem_keys = _normaliza_lista_lojas_param(lojas)

    if not ordem_keys:
        raise ValueError("Nenhuma loja informada/derivada do parâmetro.")

    # 4) Monta listas NA ORDEM selecionada, pegando a primeira linha quando houver duplicidade
    df_idx = df.set_index("_LOJA_KEY")
    lojas_out, codigos, usuarios, senhas = [], [], [], []

    lojas_inexistentes = []
    for key in ordem_keys:
        if key not in df_idx.index:
            lojas_inexistentes.append(key)
            continue

        row = df_idx.loc[key]
        if isinstance(row, pd.DataFrame):
            row = row.iloc[0]  # pega a primeira ocorrência

        lojas_out.append(_norm_loja(row["LOJAS"]))        # mantém padronização usada no resto do fluxo
        codigos.append(str(row["CODIGOS"]).strip())
        usuarios.append(str(row["USUARIOS"]).strip())
        senhas.append(str(row["SENHAS"]).strip())

    if lojas_inexistentes:
        # falha explícita para evitar executar sem todas as lojas pedidas
        raise ValueError(f"Lojas não encontradas no Excel (use nomes válidos ou equivalentes normalizados): {lojas_inexistentes}")

    return {"LOJAS": lojas_out, "CODIGOS": codigos, "USUARIOS": usuarios, "SENHAS": senhas}

CRITICAS_PDF_XLSX = {"JUAZEIRO", "CRATO", "NOVA_ONDA_ARACATI"}

def verificar_arquivos(pasta_downloads: str | os.PathLike, lojas: list[str]) -> dict:
    """
    Varre a pasta e retorna o status por loja.
    Retorna dict com chaves:
      - missing_zip: set(lojas)
      - need_extract_from_zip: set(lojas)
      - missing_pdf_or_xlsx: set(lojas)
    """
    base = P(pasta_downloads)
    status = {
        "missing_zip": set(),
        "need_extract_from_zip": set(),
        "missing_pdf_or_xlsx": set(),
    }

    for loja in lojas:
        loja_u = str(loja).upper()
        zip_path   = base / f"{loja_u}.zip"
        txt_path   = base / f"{loja_u}.txt"
        pdf_path   = base / f"{loja_u}.pdf"
        xlsx_path  = base / f"extracao_pdf_{loja_u}.xlsx"

        has_zip  = zip_path.exists()
        has_txt  = txt_path.exists()
        has_pdf  = pdf_path.exists()
        has_xlsx = xlsx_path.exists()

        if not has_zip:
            status["missing_zip"].add(loja_u)
        elif has_zip and not has_txt:
            status["need_extract_from_zip"].add(loja_u)

        if loja_u in CRITICAS_PDF_XLSX and (not has_pdf or not has_xlsx):
            status["missing_pdf_or_xlsx"].add(loja_u)

    return status


def baixa_arquivos_cnh_honda_main(lojas: str, *, retries: int = 0, max_retries: int = 1):
    hoje = datetime.now()

    path_file_log = r'\\172.17.67.14\findev$\Automação - CNH\Baixa de Arquivos\Logs\\'
    arquivo_log = f'logs_baixas-{hoje}.log'.replace('/', '-').replace(':', '_')
    caminho = os.path.join(path_file_log, arquivo_log)

    logging.basicConfig(
        filename=caminho,
        level=logging.ERROR,
        format="%(message)s"
    )

    # Marca o início da execução (para métricas)
    start_time = time()

    path = Path()

    try:
        driver, wdw, PASTA_DOWNLOADS = _ensure_driver()
        PASTA_DOWNLOADS = PASTA_DOWNLOADS / r'Automação - CNH\Baixa de Arquivos\Arquivos Baixados'
    except Exception as e:
        return False, f'Erro ao criar o webdriver.\nDescrição: {str(e)}'

    # --- monta lista das lojas realmente processadas (em maiúsculas) ---
    try:
        usuarios = get_all_users(lojas)
        lojas_processadas = [u.nome_loja.upper() for u in usuarios]
    except Exception as e:
        return False, f'Erro ao buscar os usuários.\nDescrição: {str(e)}'

    try:
        driver.get(path.Url.url)
    except Exception as e:
        return False, f'Erro ao abrir o site.\nDescrição: {str(e)}'

    for user in usuarios:
        pular_loja = False
        try:
            # ===============================
            #             LOGIN
            # ===============================
            try:
                espera_personalizada(inicio_random=1,fim_random=3)
                driver.find_element(By.CSS_SELECTOR, path.Login.campo_code).send_keys(user.codigo)
                espera_personalizada(inicio_random=1,fim_random=3)
                driver.find_element(By.CSS_SELECTOR, path.Login.campo_user).send_keys(user.usuario)
                espera_personalizada(inicio_random=1,fim_random=3)
                driver.find_element(By.CSS_SELECTOR, path.Login.campo_password).send_keys(user.senha)
                espera_personalizada(inicio_random=1,fim_random=3)

                submit = wdw.until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, path.Login.btn_entrar))
                )
                submit.click()
            except Exception as e:
                logging.error(f'Erro ao fazer o login.\nDescrição: {str(e)}\n{'-'*60}')
                driver.get(path.Url.url)
                continue

            try: 
                prosseguir = wdw.until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, path.Login.btn_prosseguir))
                )

                prosseguir.click()
            except Exception as e:
                logging.error(f'Erro ao quebrar o captcha.\nDescrição: {str(e)}\n{'-'*60}')
                driver.get(path.Url.url)
                continue

            # ===============================
            #       NAVEGAÇÃO NO MENU
            # ===============================
            espera_personalizada()

            try:
                btn_mensagem = WebDriverWait(driver, timeout=15).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, path.MenuPrincipal.btn_mensagem))
                )

                btn_mensagem.click()
            except:
                pass

            try:
                clica_na_aba(wdw, path.MenuPrincipal.aba_consorcio, 'consórcio')  # Consórcio
                espera_personalizada(inicio_random=1, fim_random=3)
                clica_na_aba(wdw, path.MenuPrincipal.aba_formularios_download, 'formulários e download')

                espera_personalizada(inicio_random=1, fim_random=3)
                clica_na_aba(wdw, path.MenuPrincipal.aba_consorcio_side_menu, 'consórcio - formulários/download')

                espera_personalizada(inicio_random=1, fim_random=3)
                clica_na_aba(wdw, path.MenuPrincipal.aba_solicitacao_carga, 'baixar informações')

                # ===============================
                #        ACESSO AO IFRAME
                # ===============================
                espera_personalizada(
                    lambda: driver.switch_to.default_content(),
                    lambda: driver.switch_to.frame(0),
                    inicio_random=2,
                    fim_random=4
                )
            except Exception as e:
                logging.error(f'Erro ao acessar o iframe.\nDescrição: {str(e)}\n{'-'*60}')
                driver.get(path.Url.url)
                continue

            for i in range(1,11):
                try:
                    id_link, texto_span = pega_texto_elemento(driver, path.Frame.id_link_ccc, i)
                except Exception as e:
                    logging.error(f'Erro pegar o texto do link do CCC.\nDescrição: {str(e)}\n{'-'*60}')
                    pular_loja = True
                    break
                
                if texto_span.strip().upper() == 'CONTA CORRENTE CONCESSIONARIA CRG':
                    driver.find_element(By.CSS_SELECTOR, f'{id_link} a').click()
                    
                    seleciona_uma_aba_do_navegador(driver, -1)
                    driver.maximize_window()

                    clicar_pelo_atributo(driver, 'value', 'ok', path.Frame.btn_baixar)

                    sleep(random.randint(2, 4))
                    driver.close()
                    seleciona_uma_aba_do_navegador(driver, 0, eh_para_fechar=True)
                    driver.switch_to.frame(0)

                    try:
                        sleep(random.randint(2, 4))
                        troca_nome_arquivo(PASTA_DOWNLOADS, f"{user.nome_loja}.zip")
                    except Exception as e:
                        logging.error(f'Erro ao trocar o nome do arquivo zip.\nDescrição: {str(e)}\n{'-'*60}')
                        pular_loja = True
                        break

                    try:
                        sleep(random.randint(2, 4))
                        caminho_zip = os.path.join(PASTA_DOWNLOADS, f"{user.nome_loja}.zip")
                        prefixo_zip = r"hda0334_new\d$\wwwhondaihs\internet\dwnhsfzip"
                        extrair_arquivo_alvo(caminho_zip, PASTA_DOWNLOADS, nome_arquivo=None, prefixo=prefixo_zip)
                    except Exception as e:
                        logging.error(f'Erro ao extrair o arquivo zip.\nDescrição: {str(e)}\n{'-'*60}')
                        pular_loja = True
                        break

                    try:
                        sleep(random.randint(2, 4))
                        troca_nome_arquivo(PASTA_DOWNLOADS, f"{user.nome_loja}.txt")
                    except Exception as e:
                        logging.error(f'Erro ao trocar o nome do arquivo txt.\nDescrição: {str(e)}\n{'-'*60}')
                        pular_loja = True
                        break

                    break
                
            if pular_loja:
                driver.get(path.Url.url)
                continue

            if user.nome_loja in ['JUAZEIRO', 'CRATO', 'NOVA_ONDA_ARACATI']:
                try:
                    driver.switch_to.default_content()
                    clica_na_aba(wdw, path.MenuPrincipal.inicio, 'inicio')
                    
                    espera_personalizada(inicio_random=1, fim_random=3)
                    clica_na_aba(wdw, path.MenuPrincipal.aba_consorcio, 'consórcio')  # Consórcio
                    espera_personalizada(inicio_random=1, fim_random=3)
                    clica_na_aba(wdw, path.MenuPrincipal.aba_formularios_download, 'financeiro')

                    espera_personalizada(inicio_random=1, fim_random=3)
                    clica_na_aba(wdw, path.MenuPrincipal.aba_consorcio_side_menu, 'consórcio - financeiro')

                    espera_personalizada(inicio_random=1, fim_random=3)
                    clica_na_aba(wdw, path.MenuPrincipal.aba_solicitacao_carga, 'conta corrente concessionária')

                    # ===============================
                    #        ACESSO AO JANELA
                    # ===============================
                    seleciona_uma_aba_do_navegador(driver, -1)
                except Exception as e:
                    logging.error(f'Erro ao acessar a tela para baixar o pdf.\nDescrição: {str(e)}\n{'-'*60}')
                    seleciona_uma_aba_do_navegador(driver, -1, eh_para_fechar=True)
                    driver.close()
                    seleciona_uma_aba_do_navegador(driver, 0, eh_para_fechar=True)
                    driver.get(path.Url.url)
                    continue

                # --- Define data de ontem ---
                ontem = datetime.now() - timedelta(days=1)
                if ontem.weekday() == 6:  # se for domingo, volta 2 dias (sexta-feira)
                    ontem = ontem - timedelta(days=2)
                data_ontem = ontem.strftime("%d/%m/%y")

                sleep(random.randint(2, 4))

                try:
                    # --- Preenche período da consulta ---
                    preencher_campo(driver, path.Janela.input_inicio, data_ontem)
                    preencher_campo(driver, path.Janela.input_fim, data_ontem)

                    espera_personalizada(
                        lambda: driver.find_element(By.CSS_SELECTOR, 'select').click(),
                        inicio_random=1,
                        fim_random=3
                    )

                    clica_na_aba(wdw, 'select option', 'outros')

                    clicar_pelo_atributo(driver, 'value', 'confirmar', path.Janela.btn_confirmar)
                    
                    linha_tabela = wdw.until(
                        EC.presence_of_element_located((By.ID, "Grid2ContainerRow_0001"))
                    )
                    sleep(1)
                except Exception as e:
                    logging.error(f'Erro baixar o arquivo pdf.\nDescrição: {str(e)}\n{'-'*60}')
                    seleciona_uma_aba_do_navegador(driver, -1, eh_para_fechar=True)
                    driver.close()
                    seleciona_uma_aba_do_navegador(driver, 0, eh_para_fechar=True)
                    driver.get(path.Url.url)
                    continue

                try:
                    clicar_pelo_atributo(driver, 'value', 'imprimir', path.Janela.btn_imprimir)

                    troca_nome_arquivo(PASTA_DOWNLOADS, f"{user.nome_loja}.pdf")
                except Exception as e:
                    logging.error(f'Erro ao trocar o nome do pdf.\nDescrição: {str(e)}\n{'-'*60}')
                    for _ in range(2):
                        seleciona_uma_aba_do_navegador(driver, -1, eh_para_fechar=True)
                        driver.close()
                    seleciona_uma_aba_do_navegador(driver, 0, eh_para_fechar=True)
                    driver.get(path.Url.url)
                    continue

                for _ in range(2):
                    seleciona_uma_aba_do_navegador(driver, -1, eh_para_fechar=True)
                    driver.close()
                
                seleciona_uma_aba_do_navegador(driver, 0, eh_para_fechar=True)
                driver.switch_to.frame(0)

                try:
                    extract_text_pdfplumber(nome_loja=user.nome_loja)
                except Exception as e:
                    logging.error(f'Erro ao extrair os dados do pdf para o excel.\nDescrição: {str(e)}\n{'-'*60}')
                    driver.get(path.Url.url)
                    continue

            sleep(random.randint(2, 4))
            # ===============================
            #          SAIR DA LOJA
            # ===============================
            sair_ihs(driver, path.Logout.btn_sair)

        except Exception as e:
            logging.error(f'Erro ao solicitar a carga da loja {user.nome_loja}.\nDescrição: {str(e)}')
            
            abas = espera_personalizada(
                lambda: driver.window_handles,
                retorno=True
            )
            
            if len(abas) > 1:
                driver.close()
                seleciona_uma_aba_do_navegador(driver, 0, eh_para_fechar=True)
                
            driver.get(path.Url.url)
            continue

    # ===============================
    #   VERIFICAÇÃO PÓS-PROCESSO
    # ===============================
    prefixo_zip = r"hda0334_new\d$\wwwhondaihs\internet\dwnhsfzip"  # o mesmo que você usa no fluxo

    status = verificar_arquivos(PASTA_DOWNLOADS, lojas_processadas)
    missing_zip = status["missing_zip"]
    need_extract = status["need_extract_from_zip"]
    missing_pdf_or_xlsx = status["missing_pdf_or_xlsx"]

    # 1) Se houver .zip presente mas faltando .txt → extrai agora (sem refazer login)
    for loja in need_extract:
        try:
            caminho_zip = os.path.join(PASTA_DOWNLOADS, f"{loja}.zip")
            extrair_arquivo_alvo(caminho_zip, PASTA_DOWNLOADS, nome_arquivo=None, prefixo=prefixo_zip)
            # caso sua extração gere nome diferente, garanta o rename:
            try:
                troca_nome_arquivo(PASTA_DOWNLOADS, f"{loja}.txt")
            except Exception:
                pass
        except Exception as e:
            logging.error(f"[PÓS] Erro ao extrair zip pendente da loja {loja}: {e}")

    # 2) Recalcula após extração
    status2 = verificar_arquivos(PASTA_DOWNLOADS, lojas_processadas)
    missing_zip = status2["missing_zip"]
    missing_pdf_or_xlsx = status2["missing_pdf_or_xlsx"]

    # 3) Se ainda faltar ZIP (nada baixado) OU faltar PDF/XLSX (para lojas críticas)
    lojas_para_refazer = set()
    lojas_para_refazer |= missing_zip
    lojas_para_refazer |= missing_pdf_or_xlsx

    if lojas_para_refazer and retries < max_retries:
        # Refaz SOMENTE as lojas pendentes
        lojas_refazer_str = ",".join(sorted(lojas_para_refazer))
        logging.error(f"[PÓS] Reexecutando para lojas pendentes: {lojas_refazer_str} (tentativa {retries+1}/{max_retries})")
        # chamada recursiva controlada
        return baixa_arquivos_cnh_honda_main(lojas_refazer_str, retries=retries+1, max_retries=max_retries)

    # 4) Se ainda assim sobrou pendência, registra e finaliza com alerta
    if lojas_para_refazer:
        pend = ",".join(sorted(lojas_para_refazer))
        logging.error(f"[PÓS] Pendências após {max_retries+1} tentativa(s): {pend}")
        return False, f"Concluído com pendências: {pend}. Verifique credenciais/rede/antivírus."

    return True, 'A baixa dos arquivos da Honda foi efetuada corretamente.'


